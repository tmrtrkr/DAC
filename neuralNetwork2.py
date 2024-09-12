import os
os.environ['TF_ENABLE_ONEDNN_OPTS'] = '0'

import tensorflow.python.keras.layers as layers
import tensorflow.python.keras.metrics as metrics
from tensorflow.python.keras import optimizers
from tensorflow.python.keras import models
from sklearn.preprocessing import StandardScaler
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from tensorflow.python.keras.optimizer_v2 import adam as adam_v2
from openpyxl import Workbook
from openpyxl import load_workbook



from tensorflow.python.keras.engine import data_adapter

def _is_distributed_dataset(ds):
    return isinstance(ds, data_adapter.input_lib.DistributedDatasetSpec)

data_adapter._is_distributed_dataset = _is_distributed_dataset





train_df = pd.read_csv('train_sample.csv')
test_df = pd.read_csv('test_sample.csv')
print("1")
# Özellikleri ve hedef sütunu ayır
features = ['platform', 'has_ios_att_permission', 'device_category', 'release_date', 'gdp', 
            'device_brand_advan', 'device_brand_aiprotablet', 'device_brand_alcatel', 'device_brand_apple', 
            'device_brand_azumi', 'device_brand_benco', 'device_brand_blackview', 'device_brand_blu', 
            'device_brand_cheers', 'device_brand_cherry mobile', 'device_brand_cubot', 'device_brand_docomo', 
            'device_brand_doogee', 'device_brand_evercoss', 'device_brand_evertek', 'device_brand_fairphone', 
            'device_brand_gol', 'device_brand_google', 'device_brand_guophone', 'device_brand_hafury', 
            'device_brand_honor', 'device_brand_huawei', 'device_brand_infinix', 'device_brand_iocean', 
            'device_brand_k-touch', 'device_brand_karbonn', 'device_brand_kassel', 'device_brand_lava', 
            'device_brand_leagoo', 'device_brand_lenovo', 'device_brand_lg', 'device_brand_meitu', 
            'device_brand_meizu', 'device_brand_mobell', 'device_brand_motorola', 'device_brand_mozilla', 
            'device_brand_nokia', 'device_brand_nuu', 'device_brand_oale', 'device_brand_omes', 
            'device_brand_omix', 'device_brand_oneplus', 'device_brand_oppo', 'device_brand_orbic', 
            'device_brand_oukitel', 'device_brand_phonemax', 'device_brand_poco', 'device_brand_pritom', 
            'device_brand_realme', 'device_brand_samsung', 'device_brand_t-mobile', 'device_brand_tagital', 
            'device_brand_tcl', 'device_brand_teclast', 'device_brand_tecno', 'device_brand_true', 
            'device_brand_ulefone', 'device_brand_umidigi', 'device_brand_vivo', 'device_brand_whoop', 
            'device_brand_wiko', 'device_brand_winnovo', 'device_brand_xgody', 'device_brand_xiaomi', 
            'device_brand_xmobile', 'device_brand_yestel', 'device_brand_zte', 'ad_network_Cross_sale', 
            'ad_network_Facebook Ads', 'ad_network_applovin_int', 'ad_network_googleadwords_int', 
            'ad_network_ironsource_int', 'ad_network_organic', 'ad_network_restricted', 'ad_network_unityads_int', 
            'total_retention', 'total_level_advanced', 'total_level_duration', 'total_ad_revenue', 
            'total_iap_revenue', 'first_prediction']
target = 'TARGET'
print("2")
# Veriyi ölçeklendir
scaler = StandardScaler()
X = train_df[features].values
y = train_df[target].values  # Hedef sütunu eğitim setinden al
X_scaled = scaler.fit_transform(X)

# Train/test ayrımı
split_idx = int(0.8 * len(X_scaled))
X_train, X_test = X_scaled[:split_idx], X_scaled[split_idx:]
y_train, y_test = y[:split_idx], y[split_idx:]

X_train = np.array(X_train)
y_train = np.array(y_train)
X_test = np.array(X_test)
y_test = np.array(y_test)
print("3")
# Hiperparametre havuzu
learning_rates = [0.001, 0.01, 0.1]
batch_sizes = [32, 64, 128]
epochs = [50, 100, 150]
print("4")
# Neural network modeli
def build_model(learning_rate):
    model = models.Sequential([
        layers.Dense(128, activation='relu', input_shape=(X_train.shape[1],)),
        layers.Dense(64, activation='relu'),
        layers.Dense(32, activation='relu'),
        layers.Dense(1)  # Tek bir çıktı katmanı, regresyon için
    ])
    print("5")
    optimizer = adam_v2.Adam(learning_rate=learning_rate)
    model.compile(optimizer=optimizer, loss='mean_squared_error', metrics=[metrics.RootMeanSquaredError()])
    print("6")
    return model

def save_to_excel(filename, learning_rate, batch_size, epochs, rmse_value):
    data = {'Learning Rate': [learning_rate], 
            'Batch Size': [batch_size], 
            'Epochs': [epochs], 
            'RMSE': [rmse_value]}
    df_to_save = pd.DataFrame(data)

    try:
        # Mevcut Excel dosyasını oku
        reader = pd.read_excel(filename, sheet_name='results', engine='openpyxl')
        book = load_workbook(filename)

        # `with` bloğu kullanarak writer oluştur ve mevcut dosyayı güncelle
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Mevcut çalışma sayfalarını writer'a ekleyelim
            writer.sheets.update({ws.title: ws for ws in book.worksheets})

            # Verileri mevcut verilerin altına ekle
            df_to_save.to_excel(writer, sheet_name='results', index=False, header=False, startrow=len(reader)+1)
    
    except FileNotFoundError:
        # Eğer dosya yoksa, yeni bir dosya oluştur
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df_to_save.to_excel(writer, sheet_name='results', index=False, header=True)

    print("Results saved to", filename)

    
# Modeli eğit ve tahmin et
for lr in learning_rates:
    for bs in batch_sizes:
        for ep in epochs:
            # Modeli oluştur ve eğit
            print("10")
            model = build_model(lr)
            model.fit(X_train, y_train, epochs=ep, batch_size=bs, verbose=0)
            print("11")
            # Tahmin yap ve RMSE hesapla
            predictions = model.predict(X_test)
            rmse_value = np.sqrt(np.mean((predictions - y_test) ** 2))
            
            # Sonuçları Excel'e kaydet (sadece genel performans)
            save_to_excel('predictions_metrics.xlsx', lr, bs, ep, rmse_value)
            print(f"Model trained with LR={lr}, Batch Size={bs}, Epochs={ep}")
            print(f"RMSE: {rmse_value}")

# Tüm işlemler tamamlandığında
print("All model training and predictions are complete.")

